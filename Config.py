  ###################################### SimplifAI ##########################################
##                              File for Config Manipulation                               ##
##                                     Author: Erwan                                       ##
##                                   Date: 2023-03-20                                      ##
##                                     Version: 1.0                                        ##
##                                Python Version: 3.10.6                                   ##
  ###################################### SimplifAI ##########################################

import pprint
import openpyxl

PATH = "./Config/"
LANGUAGE = "en-US"
MAX_TOKENS = "100"
PROMPT = "(Act as you were a robot talking, generate a text that adapt to the user, your name is SimplifAI, Qestion is the one you need to answer, The context is to help you find an appropriated answer for the situation) do a long explained answer."
CONTEXT = "Home"

def CreateExceFile():
    wb = openpyxl.Workbook()

    # Creating a new sheet
    sheet = wb.active
    sheet.title = "Config"

    # Creating the first row
    sheet["A1"] = "Variable Name"
    sheet["B1"] = "Value"
    sheet["C1"] = "Description"

    # Creating the first column
    sheet["A2"] = "Path"
    sheet["A3"] = "Language"
    sheet["A4"] = "Max Tokens"
    sheet["A5"] = "Prompt"
    sheet["A6"] = "Context"

    # Creating the second column
    sheet["B2"] = PATH
    sheet["B3"] = LANGUAGE
    sheet["B4"] = MAX_TOKENS
    sheet["B5"] = PROMPT
    sheet["B6"] = CONTEXT

    # Creating the third column
    sheet["C2"] = "Path to the folder where the files will be saved"
    sheet["C3"] = "Language of the text to generate"
    sheet["C4"] = "Maximum number of tokens to generate, for chat gpt"
    sheet["C5"] = "Prompt to generate the text"
    sheet["C6"] = "where the action take place"


    # Saving the workbook to     the path
    wb = wb.save(PATH + "Config" + LANGUAGE + ".xlsx")


def FillDictionary( Path, Language, MaxTokens, Prompt, Context ):
    Dictionary = {}
    Dictionary["Path"] = Path
    Dictionary["Language"] = Language
    Dictionary["MaxTokens"] = MaxTokens
    Dictionary["Prompt"] = Prompt
    Dictionary["Context"] = Context

    print("Dictionary filled")
    return Dictionary

def ModifyRow( Path, Language, MaxTokens, Prompt, Context ):
    wb = openpyxl.load_workbook( Path + "Config" + Language + ".xlsx" )
    sheet = wb["Config"]

    sheet["B4"] = MaxTokens

    wb.save( Path + "Config" + Language + ".xlsx" )

def ReadConfigFile( Path, Language ):
    wb = openpyxl.load_workbook( Path + "Config" + Language + ".xlsx" )
    sheet = wb["Config"]

    Path = sheet["B2"].value
    Language = sheet["B3"].value
    MaxTokens = sheet["B4"].value
    Prompt = sheet["B5"].value
    Context = sheet["B6"].value

    return FillDictionary( Path, Language, MaxTokens, Prompt, Context)

def ReadKeywords( Path):
    wb = openpyxl.load_workbook( Path + "Config.xlsx" )
    sheet = wb["Config"]

    # create dict[language][keyword]
    KeywordsActivated = {}
    KeywordsDeactivated = {}
    languages = []

    # Read the first row
    for i in range(2, sheet.max_row + 1):
        # get first cell
        cell = sheet.cell(row=i, column=1)
        cell2 = sheet.cell(row=i, column=2)
        language = sheet.cell(row=i, column=3)

        # cell is a string separated by a , and a space
        # we need to split it to get the list of keywords
        Keywords1 = cell.value.split(", ")
        Keywords2 = cell2.value.split(", ")

        # add it keywords1 and cell3 and keywords2 and cell3 to the keywords activated and deactivated
        for keyword in Keywords1:
            if language.value not in KeywordsActivated:
                KeywordsActivated[language.value] = []
            KeywordsActivated[language.value].append(keyword)
        for keyword in Keywords2:
            if language.value not in KeywordsDeactivated:
                KeywordsDeactivated[language.value] = []
            KeywordsDeactivated[language.value].append(keyword)

        languages.append(language.value)

    # display language of one of the keywords
    # print(KeywordsActivated["en-US"][1])

    # print(KeywordsActivated)
    # print(KeywordsDeactivated)
    return KeywordsActivated, KeywordsDeactivated, languages

def DisplayDictionary( Dictionary ):
    # display the dictionary with  the lib pprint
    pprint.pprint( Dictionary )

# def main():
#     CreateExceFile()
#     Dictionary = ReadConfigFile(PATH, LANGUAGE)
#     DisplayDictionary(Dictionary)
#     # modifyrow(PATH, LANGUAGE, "200", PROMPT, CONTEXT)
#     Dictionary = ReadConfigFile(PATH, LANGUAGE)
#     DisplayDictionary(Dictionary)

# if __name__ == "__main__":
#     main()
